# -*- coding:utf-8 _*-
from tools.project_path import ProjectPath
from openpyxl import load_workbook
import time
from test_data.helper import globals_data

class Do_excel:

    def get_data(self, filename, sheet_name):
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        test_data = []
        for i in range(2, sheet.max_row+1):
            data = {}
            data['case_id'] = sheet.cell(i, 1).value
            data['url'] = sheet.cell(i, 2).value
            data['data'] = sheet.cell(i, 3).value
            # replace_data = ['NONE','TOKEN']
            # for j in replace_data:
            #     if data['data'].find('$'+j) != -1:
            #         data['data'] = data['data'].replace('$'+j, globals_data[j])

            data['title'] = sheet.cell(i, 4).value
            data['http_method'] = sheet.cell(i, 5).value
            data['excepted'] = sheet.cell(i, 6).value
            data['headers'] = sheet.cell(i, 9).value
            data['result'] = sheet.cell(i, 7).value
            test_data.append(data)
        return test_data

    def write_back(self, filename, sheet_name, i, value, TestResult):
        """写回数据到Excel中"""
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value = value
        sheet.cell(i, 8).value = TestResult
        wb.save(filename)



if __name__ == '__main__':
    data = Do_excel().get_data(ProjectPath.test_data_path, 'Sheet1')
    print(data)

    test = {'is_ziti':'0','lat':'24.61013293104972','lng':'118.0467790667441','products':' [{"product_id" : "363863","spec_id" : "","num" : 2,"ingredient" : [],"specification" : [],"package_id" : "0"}]','shop_id':'1294','timestamp':'','ver':'3.1.4',}